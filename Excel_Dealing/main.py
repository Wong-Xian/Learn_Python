import openpyxl

if __name__ == '__main__':

    excel_File = openpyxl.load_workbook('薛城完整数据自动化整理.xlsx')
    sheet_Of_file = excel_File.worksheets[0]
    #
    # # for row_r in range(3, 60):
    #
    # row_r = 3345
    # row_c = 3
    #
    # while row_r < 4477:
    #     ref = sheet_Of_file.cell(row_r, 1).value  # 参考值
    #     # print('F', row_r, "的值是", ref)
    #
    #     to_Compare = sheet_Of_file.cell(row_c, 1).value  # 要对比的值
    #     # print('A', row_c, "的值是", to_Compare)
    #
    #     if ref == to_Compare:
    #         print('第', row_c - 2, '行相等')
    #         row_r += 1
    #         row_c += 1
    #     else:
    #         print('第', row_c - 2, '行不相等')
    #         sheet_Of_file.delete_rows(row_c)
    #         row_r -= 1
    #         print(row_r)
    #         print('已删除第', row_c, '行')
    #
    # print('保存文件到try1.xlsx')
    # excel_File.save('try1.xlsx')

    # data = sheet_Of_file.cell(1, 1).value
    # print(data)

    row = 3

    while row < (90 - 3) / 3 + 3:
        print(row)
        sheet_Of_file.delete_rows(row + 1)
        sheet_Of_file.delete_rows(row + 1)
        row += 1

    excel_File.save('薛城完整数据自动化整理2.xlsx')
